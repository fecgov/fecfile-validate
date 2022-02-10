"""Convert the FEC validation Excel spreadsheet into JSON schema documents.

This is a utility script to create *starter* JSON schema document templates
from the FEC validation Excel spreadsheet found in the FEC VenPak zip file
that can be downloaded from:

https://www.fec.gov/help-candidates-and-committees/filing-reports/fecfile-software/

The JSON schema standard can be found here:

http://json-schema.org/

Note: Schema properties prefixed with "fec_" are not part of the JSON schema
standard and are specific to the FEC data.
"""

from enum import Enum
import openpyxl
import json
import argparse
import os

parser = argparse.ArgumentParser(description='Convert the FEC validation Excel'
                                 ' spreadsheet into JSON schema documents.')
parser.add_argument('excel_filename', help='an excel filename that will be'
                    ' parsed to generate JSON schema docs')
parser.add_argument('--sheets-to-generate', help='a json file containing an'
                    ' array of sheet names to be parsed from the excel file')
parser.add_argument('--version')
args = parser.parse_args()
EXCEL_FILENAME = args.excel_filename or \
    "Form_3X_Receipts_Vendor_10.20.2020.xlsx"
SCHEMA_ID_PREFIX = ("https://github.com/fecgov/fecfile-validate/blob/"
                    "main/schema")
VERSION = args.version or "v0.0.0.0"
SHEETS_TO_SKIP = ['All receipts', 'Version 8.3', 'SUMMARY OF CHANGES',
                  "All Schedule A Transactions", "ScheduleC", "Schedule C1",
                  "Scedule C2"]

# Column postions of fields in the spreadsheet row array


class Columns(Enum):
    COL_SEQ = 0
    FIELD_DESCRIPTION = 1
    TYPE = 2
    REQUIRED = 3
    AUTO_POPULATE = 4
    SAMPLE_DATA = 5
    VALUE_REFERENCE = 6
    RULE_REFERENCE = 7
    FIELD_FORM_ASSOCIATION = 8

    def get(self, row, has_autopopulate):
        index = self.value if has_autopopulate or \
            self.value <= 3 else self.value - 1
        value = row[index] if index < len(row) else None
        return value.strip() if isinstance(value, str) else value

def convert_row_to_property(row, sheet_has_autopopulate):# noqa
    """Take a row from the spreadsheet and convert it into a schema object.

    Args:
        row (array): Single row from the spreadsheet

    Returns:
        token (string): The token key identifier for this property in
                        the schema.
        prop (dict): The property object to add to the schema.
    """
    prop = {}
    spec = {}
    for col in Columns:
        if col != Columns.AUTO_POPULATE or sheet_has_autopopulate:
            spec[col.name] = col.get(row, sheet_has_autopopulate)

    title = spec.get(Columns.FIELD_DESCRIPTION.name)
    field_type = spec.get(Columns.TYPE.name)
    required = spec.get(Columns.REQUIRED.name)
    sample_data = spec.get(Columns.SAMPLE_DATA.name)
    rule_ref = spec.get(Columns.RULE_REFERENCE.name)

    token = title.replace("\n", "_").replace(" ", "_").replace(".", "")\
        .replace("(", "").replace(")", "").replace("/", "_")\
        .replace("__", "_").lower()
    # Prepend tokens that start with a number (presumed to be a line number)
    # with capital letter "L".
    if token[0].isdigit():
        token = 'L' + token
    prop["title"] = title
    prop["description"] = ""

    if field_type.startswith("AMT-"):
        prop["type"] = "number"
        prop["minimum"] = 0
        prop["maximum"] = int('9' * int(field_type.split('-')[1]))

    if field_type.startswith("NUM-"):
        length = field_type.split('-')[1].strip()
        prop["type"] = "string"
        prop["minLength"] = 0
        prop["maxLength"] = int(length)
        prop["pattern"] = f'^\d{{0,{length}}}$'

        

    if field_type.startswith("A/N-") or field_type.startswith("A-"):
        if field_type == "A-1" and rule_ref == "Check-box":
            prop["type"] = "boolean"
        else:
            length = field_type.split('-')[1].strip()
            prop["type"] = "string"
            prop["minLength"] = 0
            prop["maxLength"] = int(length)
            prop["pattern"] = f'^[ A-z0-9]{{0,{length}}}$'

    if sample_data:
        prop["examples"] = [sample_data]

    is_required = required and "error" in required
    is_recommended = required and "warn" in required

    prop["fec_spec"] = spec
    return (token, prop, is_required, is_recommended)


wb = openpyxl.load_workbook(EXCEL_FILENAME)
sheets_to_generate = None
if args.sheets_to_generate is not None:
    with open(os.path.join(os.getcwd(), args.sheets_to_generate), 'r') as f:
        sheets_to_generate = json.load(f)

print(sheets_to_generate)
trans_type_hits = {}
for ws in wb.worksheets:
    if ((sheets_to_generate is not None and ws.title not in sheets_to_generate)
            or ws.title in SHEETS_TO_SKIP):
        continue

    print(ws.title)
    title = ws.title.replace(' ', '')
    output_file = title + ".json"

    print(f'Parsing {output_file}...')

    sheet_has_autopopulate = ws.cell(3, 5).value is not None and \
        ws.cell(3, 5).value.strip() == 'Auto populate'
    schema_properties = {}
    required_rows = []
    recommended_rows = []
    for row in ws.iter_rows(min_row=5, max_col=8, values_only=True):
        if (not Columns.COL_SEQ.get(row, sheet_has_autopopulate)
                or Columns.COL_SEQ.get(row, sheet_has_autopopulate) == "--"
                or not Columns.FIELD_DESCRIPTION.get(row,
                                                     sheet_has_autopopulate)
                or not Columns.TYPE.get(row, sheet_has_autopopulate)
                or len(row) > 10):
            continue
        token, prop, is_required, is_recommended = \
            convert_row_to_property(row, sheet_has_autopopulate)
        if token == "transaction_type_identifier":
            trans_type_id = \
                prop.get('fec_spec', {}).get(Columns.SAMPLE_DATA.name,
                                             "") or ""
            trans_type_hits[trans_type_id] = \
                (trans_type_hits.get(trans_type_id) or 0) + 1
            if (trans_type_hits[trans_type_id] > 1 or trans_type_id == ''):
                output_file = trans_type_id + '-' + \
                    str(trans_type_hits[trans_type_id]) + '.json'
            else:
                output_file = trans_type_id + '.json'
        # Catch and mark token (i.e. spec property) clashes for manual fixing.
        if token in schema_properties:
            token = token + '-DUPLICATE'

        if is_required:
            required_rows.append(token)
        if is_recommended:
            recommended_rows.append(token)
        schema_properties[token] = prop

    schema = {
        "$schema": "https://json-schema.org/draft-07/schema#",
        "$id": f'{SCHEMA_ID_PREFIX}/{output_file}',
        "version": VERSION,
        "title": f'FEC {ws.title}',
        "description": ws.cell(1, 1).value,
        "type": "object",
        "required": required_rows,
        "fec_recommended": recommended_rows,
        "properties": schema_properties,
        "additionalProperties": False
    }
    f = open(output_file, "w")
    f.write(json.dumps(schema, indent=4))
    f.close()
    print('Done')
