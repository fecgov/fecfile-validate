"""
Verify local JSON schema files against an Excel spreadsheet.

This script checks for differences between this repo's JSON files and a spec spreadsheet.
The script will scan the local directory for a spreadsheet ending with ".xlsx" unless
the user passes in a valid  spreadsheet file name.

The JSON schema standard can be found here:
http://json-schema.org/
"""


from os import path, listdir
from openpyxl import load_workbook
import argparse
import json
import time
import re


parser = argparse.ArgumentParser(
    description=""
    "This script checks for differences between this "
    "repo's JSON files and a spec spreadsheet.\n"
    "The script will scan the local directory for a "
    'spreadsheet ending with ".xlsx" unless the user\n'
    "passes in a valid  spreadsheet file name.\n\n"
)
parser.add_argument(
    "excel_filename",
    default=None,
    nargs="?",  # Allows for 0 or 1 filenames to be specified
    help="an excel filename that will be parsed to generate JSON schema docs",
)
parser.add_argument(
    "-v", "--verbose", help="record and print minor errors", action="store_true"
)
parser.add_argument(
    "-d",
    "--debug",
    help="Prints the names of sheets and fields as the script works",
    action="store_true",
)
args = parser.parse_args()
VERBOSE = args.verbose
DEBUG = args.debug
EXCEL_FILENAME = args.excel_filename
COLUMNS = {
    "property_name": 0,
    "type": 1,
    "required": 2,
    "sample_data": 3,
    "value_reference": 4,
    "rule_reference": 5,
}


def get_filename(sheet):
    filename_overrides = {
        (
            "NATIONAL_PARTY_PARTNERSHIP_MEMOS (needs review as it is named "
            "PARTNERSHIP_NATIONAL_PARTY_MEMOS in the code)"
        ): "PARTNERSHIP_NATIONAL_PARTY_MEMOS",
        (
            "NATIONAL_PARTY_PARTNERSHIP_RECEIPTS (needs review as it is named "
            "PARTNERSHIP_NATIONAL_PARTY_RECEIPTS in the code)"
        ): "PARTNERSHIP_NATIONAL_PARTY_RECEIPTS",
    }

    filename = sheet["A2"].value
    if not filename:
        filename = get_transaction_type_identifier(sheet)
    elif filename in filename_overrides.keys():
        filename = filename_overrides[filename]

    return filename


def get_transaction_type_identifier(sheet):
    column_overrides = {"OFFSET_TO_OPERATING_EXPENDITURE": "E"}

    tti_row_range = range(5, 11)
    for row in tti_row_range:
        column = "F"
        if sheet.title in column_overrides.keys():
            column = column_overrides[sheet.title]
        if sheet[f"A{row}"].value == "TRANSACTION TYPE IDENTIFIER":
            return sheet[f"{column}{row}"].value

    return ""


def get_schema_fields(sheet):
    name_overrides = {
        "contribution_purpose_description": "contribution_purpose_descrip",
        "memo_text/description": "memo_text_description",
        "contributor_organization": "contributor_organization_name",
        "contributor_street__1": "contributor_street_1",
        "contributor_street__2": "contributor_street_2",
    }

    skipped_fields = [
        "receipt_description",
    ]

    fields = {}
    row = None
    for r in range(1, sheet.max_row):
        field_description = sheet[f"A{r}"].value
        if field_description == "FORM TYPE":
            row = r
            break

    if row is None:
        print("    FORM TYPE row not found", sheet.title)

    if DEBUG:
        print("    Determining the schema spreadsheet's fields")

    while sheet["A" + str(row)].value:
        if DEBUG and VERBOSE:
            print(f"        Parsing row {row}")

        field_name_raw = sheet["A" + str(row)].value.lower()
        field_name = "_".join(field_name_raw.split(" "))
        if field_name in name_overrides:
            field_name = name_overrides[field_name]

        if field_name[-1] == "_":
            field_name = field_name[:-1]

        if field_name not in skipped_fields:
            fields[field_name] = row

        row += 1

    return fields


def get_schema_property(schema, field_name, property, is_fec_spec=False):
    field = schema["properties"][field_name]
    if is_fec_spec:
        field = field["fec_spec"]

    if property in field.keys():
        return field[property]
    else:
        return None


def check_transaction_type_identifier(row, schema, field_name):
    errors = []
    if field_name != "transaction_type_identifier":
        return errors

    sheet_tti = row[COLUMNS["rule_reference"]].value
    if not sheet_tti:
        errors.append(
            f"    Error: {field_name} - The sheet's Transaction Type "
            + "Identifier is blank"
        )
        return errors

    if "\n" in sheet_tti:
        sheet_tti = sheet_tti.split("\n")
        return check_multi_tti(schema, field_name, sheet_tti)
    else:
        return check_single_tti(schema, field_name, sheet_tti)


def check_single_tti(schema, field_name, sheet_tti):
    errors = []
    json_tti = get_schema_property(schema, field_name, "const")
    if not json_tti:
        errors.append(
            f'    Error: {field_name} - Schema has a TTI of "{sheet_tti}"'
            " while the JSON does not have a constant value"
        )
    elif json_tti != sheet_tti:
        errors.append(
            f'    Error: {field_name} - Schema has a TTI of "{sheet_tti}"'
            f' while the JSON has "{json_tti}"'
        )
    return errors


def check_multi_tti(schema, field_name, sheet_tti):
    errors = []
    json_tti = get_schema_property(schema, field_name, "enum")
    if not json_tti:
        errors.append(
            f"    Error: {field_name} - Schema allows for TTI's of {sheet_tti}"
            " while the JSON does not have an enumerated value"
        )
    else:
        for j_tti in json_tti:
            if j_tti not in sheet_tti:
                errors.append(
                    f'    Error: {field_name} - The JSON has a TTI of "{j_tti}"'
                    " that is not present in the sheet"
                )
        for s_tti in sheet_tti:
            if s_tti not in json_tti:
                errors.append(
                    f'    Error: {field_name} - The Sheet has a TTI of "{s_tti}"'
                    " that is not present in the JSON"
                )
    return errors


def compare_type(row, schema, field_name):
    errors = []
    expected_type = row[COLUMNS["type"]].value
    actual_type = get_schema_property(schema, field_name, "TYPE", is_fec_spec=True)

    if expected_type is None:
        match = True
    else:
        # Strips the lingering spaces present in some fields of the spreadsheet
        expected_type = expected_type.strip(" ")
        match = expected_type == actual_type

    if not match:
        errors.append(
            f"    Error: {field_name} - Sheet has Type {expected_type} "
            f"and the JSON has {actual_type}"
        )

    return errors


def get_length_from_type(type):
    if not type:
        return None

    split_type = type.split("-")
    if len(split_type) <= 1:
        return None

    return int(split_type[-1].strip(" "))


def get_cpd_lengths_and_patterns(regex):
    lengths = []
    patterns = regex.split("|")
    for pattern in patterns:
        split_left = pattern.split("[")
        split_right = pattern.split("}")

        fixed = []
        if len(split_left) > 1:
            fixed.append(split_left[0])
        if len(split_right) > 1:
            fixed.append(split_right[1])

        # Right of the "[" we find the "[ -~]{min_length,max_length}$" pattern
        length_str = split_left[-1]

        # Finding the maximum length of abritrary characters in the pattern
        length_str = length_str.split(",")[1]  # Right of the comma aaaaand
        max_length = length_str.split("}")[0]  # left of the brace is the max length

        length = 0
        # Remove the ^, &, \\, (, and ) symbols since they're not part of the length
        # Count the length of escaped parens!
        for f in fixed:
            fix = f.strip("^").strip("$")
            length += fix.count("\\)") + fix.count("\\(")
            fix = fix.replace("\\", "").replace("(", "").replace(")", "")
            length += len(fix)

        lengths.append([length + int(max_length), pattern])

    return lengths


def compare_length(row, schema, field_name):
    errors = []

    # These are recurring patterns whose lengths are very hard to measure with a function
    # Instead, the fields corresponding to the keys here are considered as matching if
    # their pattern matches the value below
    fixed_patterns = {
        "filer_committee_id_number": "^(?:[PC][0-9]{8}|[HS][0-9]{1}[A-Z]{2}[0-9]{5})$",
        "contribution_date": "^[0-9]{4}-[0-9]{2}-[0-9]{2}$",
        "donor_committee_fec_id": "^(?:[PC][0-9]{8}|[HS][0-9]{1}[A-Z]{2}[0-9]{5})$",
    }

    field_type = row[COLUMNS["type"]].value
    expected_length = get_length_from_type(field_type)
    if not expected_length:
        return errors

    json_max_length = get_schema_property(schema, field_name, "maxLength")
    json_pattern_regex = get_schema_property(schema, field_name, "pattern")

    if json_max_length and json_max_length != expected_length:
        errors.append(
            f"    Error: {field_name} - Sheet has Type {field_type} but "
            f"the JSON's max_length is {json_max_length}"
        )

    if json_pattern_regex:
        if field_name == "contribution_purpose_descrip":
            lengths_and_patterns = get_cpd_lengths_and_patterns(json_pattern_regex)
            for pattern_length, pattern in lengths_and_patterns:
                if expected_length != pattern_length:
                    # Phrase the error different depending on whether or not
                    # it is a sub-pattern
                    substring = "field's pattern's"
                    if len(lengths_and_patterns) > 0:
                        substring = f"field has a sub-pattern ({pattern}) whose"

                    errors.append(
                        f"    Error: {field_name} - Sheet has Type {field_type} but "
                        f"the JSON {substring} length adds up to {pattern_length}"
                    )
        elif field_name not in fixed_patterns.keys():
            if not re.search(f"{expected_length}}}\\$$", json_pattern_regex):
                errors.append(
                    f"    Error: {field_name} - Sheet has Type {field_type} but "
                    f"the JSON field's pattern's max length is wrong "
                    f"({json_pattern_regex})"
                )
        else:
            if json_pattern_regex != fixed_patterns[field_name]:
                errors.append(
                    f"    Error: {field_name} - The JSON has a pattern of "
                    f"{json_pattern_regex} "
                    f"but the expected pattern is {fixed_patterns[field_name]}"
                )

    return errors


def check_contribution_amount(row, schema, field_name):
    errors = []

    if field_name != "contribution_amount":
        return errors

    sheet_form_type = row[COLUMNS["type"]].value
    expected_length = get_length_from_type(sheet_form_type)
    if not expected_length:
        return errors

    sheet_rule_reference = row[COLUMNS["rule_reference"]].value
    sheet_amount_negative = False
    if sheet_rule_reference is not None:
        sheet_amount_negative = "negative" in sheet_rule_reference.lower()

    json_minimum = get_schema_property(schema, field_name, "minimum")
    json_maximum = get_schema_property(schema, field_name, "maximum")
    json_exclusive_maximum = get_schema_property(schema, field_name, "exclusiveMaximum")

    if json_minimum is None:
        errors.append(
            f"    Error: {field_name} - The JSON for " f"the field has no minimum value"
        )
    elif json_minimum != 0 and len(str(json_minimum)) != expected_length:
        errors.append(
            f"    Error: {field_name} - The JSON's minimum value "
            f"is {json_minimum} when it should have a length of 12"
        )

    if not sheet_amount_negative:
        if json_maximum is None:
            errors.append(
                f"    Error: {field_name} - The JSON for the field "
                f"has no maximum value"
            )
        elif len(str(json_maximum)) != expected_length:
            errors.append(
                f"    Error: {field_name} - The JSON's maximum value "
                f"is {json_maximum} when it should have a length of 12"
            )
    else:
        if json_exclusive_maximum is None:
            errors.append(
                f"    Error: {field_name} - The JSON for "
                f"the field has no exclusiveMaximum value"
            )
        elif json_exclusive_maximum != 0:
            errors.append(
                f"    Error: {field_name} - The JSON's "
                f"exclusiveMaximum should equal 0"
            )

    return errors


def check_required(row, schema, field_name):
    sheet_required_raw = row[COLUMNS["required"]].value
    if sheet_required_raw is not None:
        sheet_required = sheet_required_raw.strip()

        if sheet_required == "X (error)":
            return check_strictly_required(schema, field_name)

        if sheet_required == "X (conditional error)":
            return check_conditionally_required(schema, field_name)

    return check_not_required(schema, field_name)


def check_strictly_required(schema, field_name):
    errors = []
    schema_required = get_schema_property(
        schema, field_name, "REQUIRED", is_fec_spec=True
    )

    if schema_required != "X (error)":
        errors.append(
            f"    Error: {field_name} - The field is required, "
            'but the JSON\'s FEC Spec does not have "X (error)"'
        )

    if field_name not in schema["required"]:
        errors.append(
            f"    Error: {field_name} - The field is required, "
            "but it's not present in the JSON's required array"
        )

    return errors


def found_in_all_of(schema, field_name):
    if "allOf" in schema.keys():
        for all_of_rule in schema["allOf"]:
            if "required" in all_of_rule["then"]:
                if field_name in all_of_rule["then"]["required"]:
                    return True
    return False


def check_conditionally_required(schema, field_name):
    errors = []
    schema_required = get_schema_property(
        schema, field_name, "REQUIRED", is_fec_spec=True
    )

    if schema_required != "X (conditional error)":
        errors.append(
            f"    Error: {field_name} - The field is conditionally required, "
            'but the JSON\'s FEC Spec does not have "X (conditional error)"'
        )

    if field_name in schema["required"]:
        errors.append(
            f"    Error: {field_name} - The field is conditionally required, "
            "but it is present in the JSON's required array"
        )

    if not found_in_all_of(schema, field_name):
        errors.append(
            f"    Error: {field_name} - The field is conditionally required, "
            "but it is not present in the JSON's allOf rules"
        )

    return errors


def check_not_required(schema, field_name):
    errors = []
    schema_required = get_schema_property(
        schema, field_name, "REQUIRED", is_fec_spec=True
    )

    if schema_required is not None:
        errors.append(
            f"    Error: {field_name} - The field is not required, "
            f'but it\'s marked as "{schema_required}" in the JSON'
        )

    if field_name in schema["required"]:
        errors.append(
            f"    Error: {field_name} - The field is not required,"
            " but it's present in the JSON's required array"
        )

    if found_in_all_of(schema, field_name):
        errors.append(
            f"    Error: {field_name} - The field is not required,"
            " but it can be set as required in the JSON's AllOf"
        )

    return errors


def compare_sample_data(row, schema, field_name):
    errors = []
    sheet_sample_data = row[COLUMNS["sample_data"]].value
    schema_sample_data = get_schema_property(schema, field_name, "SAMPLE_DATA", True)
    if sheet_sample_data != schema_sample_data:
        errors.append(
            f"    Minor Error: {field_name} - The Sheet has Sample Data of "
            f'"{sheet_sample_data}" while the JSON has "{schema_sample_data}"'
        )

    return errors


def check_form_type(row, schema, field_name):
    errors = []
    if field_name not in ["form_type", "back_reference_sched_name"]:
        return errors

    form_type = row[COLUMNS["sample_data"]].value
    schema_properties = schema["properties"][field_name]

    if "const" in schema_properties.keys():
        if form_type != schema_properties["const"]:
            errors.append(
                f"    Error: {field_name} - Sheet has Form Type "
                f'"{form_type}" while the JSON has "{schema_properties["const"]}"'
            )

    elif "enum" in schema_properties.keys():
        if form_type not in schema_properties["enum"]:
            errors.append(
                f"    Error: {field_name} - Sheet has Form Type "
                f'"{form_type}" while the JSON has "{schema_properties["enum"]}"'
            )

    return errors


def check_entity_type(row, schema, field_name):
    errors = []
    if field_name != "entity_type":
        return errors

    entity_types = row[COLUMNS["value_reference"]].value
    if not entity_types:
        errors.append(f"    Error: {field_name} - Entity Types not found in sheet")
        return errors

    if "|" in entity_types:
        return check_multiple_entity_types(schema, field_name, entity_types)
    else:
        return check_single_entity_type(schema, field_name, entity_types)


def check_single_entity_type(schema, field_name, raw_sheet_entity_type):
    errors = []
    json_entity_type = get_schema_property(schema, field_name, "const")
    if not json_entity_type:
        errors.append(
            f"    Error: {field_name} - The sheet has a single entity type"
            " but the JSON does not have a constant value"
        )
        return errors

    sheet_entity_type = raw_sheet_entity_type.replace(" Only", "").replace(" only", "")

    if sheet_entity_type != json_entity_type:
        errors.append(
            f"    Error: {field_name} - The sheet has an entity type of"
            f" {sheet_entity_type}, but the JSON has {json_entity_type}"
        )
    return errors


def check_multiple_entity_types(schema, field_name, raw_sheet_entity_types):
    errors = []
    json_entity_types = get_schema_property(schema, field_name, "enum")
    if not json_entity_types:
        errors.append(
            f"    Error: {field_name} - The sheet has a single entity type"
            " but the JSON does not have an enumerated value"
        )
        return errors

    cleaned_sheet_entity_types = raw_sheet_entity_types.replace("[", "").replace(
        "]", ""
    )
    cleaner_sheet_entity_types = cleaned_sheet_entity_types.replace(" ", "")
    sheet_entity_types = cleaner_sheet_entity_types.split("|")

    for s_entity in sheet_entity_types:
        if s_entity not in json_entity_types:
            errors.append(
                f'    Error: {field_name} - The sheet has an entity type "{s_entity}"'
                " that is not present in the JSON's entity type"
            )

    for j_entity in json_entity_types:
        if j_entity not in sheet_entity_types:
            errors.append(
                f'    Error: {field_name} - The JSON has an entity type "{j_entity}"'
                " that is not present in the Sheet's entity type"
            )

    return errors


def check_aggregation_group(row, schema, field_name):
    errors = []
    if field_name != "aggregation_group":
        return errors

    sheet_aggr_group = row[COLUMNS["rule_reference"]].value
    if not sheet_aggr_group:
        sheet_aggr_group = row[COLUMNS["value_reference"]].value

    if not sheet_aggr_group:
        errors.append(
            f"    Error: {field_name} - Cannot find aggregation group in sheet"
        )
        return errors

    if "," in sheet_aggr_group:
        return check_aggregation_group_multiple(sheet_aggr_group, schema, field_name)
    else:
        return check_aggregation_group_single(sheet_aggr_group, schema, field_name)


def clean_aggregation_group_name(aggr_group):
    sheet_aggr_group = aggr_group.replace(" ", "_")
    sheet_aggr_group = sheet_aggr_group.replace("-", "_")
    sheet_aggr_group = sheet_aggr_group.upper()
    return sheet_aggr_group


def clean_aggregation_group_names(aggr_group_field):
    aggr_group_lines = aggr_group_field.split("\n")
    aggr_groups = []
    for aggr_group_line in aggr_group_lines:
        aggr_groups.append(aggr_group_line.split(", then ")[1])

    cleaned_aggregation_group_names = []
    for aggr_group in aggr_groups:
        clean_name = clean_aggregation_group_name(aggr_group)
        cleaned_aggregation_group_names.append(clean_name)
    return cleaned_aggregation_group_names


def check_aggregation_group_single(sheet_aggr_group, schema, field_name):
    errors = []

    schema_group_name = get_schema_property(schema, field_name, "const")
    if not schema_group_name:
        errors.append(
            f"    Error: {field_name} - Cannot find aggregation group in json schema"
        )
        return errors

    sheet_group_name = ""
    if sheet_aggr_group:
        sheet_group_name = clean_aggregation_group_name(sheet_aggr_group)

    if sheet_group_name != schema_group_name:
        errors.append(
            f"    Error: {field_name} - Sheet has an (adjusted) Aggregation Group "
            f'of "{sheet_group_name}" while the JSON has "{schema_group_name}"'
        )

    return errors


def group_in_all_of(schema, group):
    group_name = group.replace("XX", "")
    if "allOf" in schema.keys():
        for all_of_rule in schema["allOf"]:
            if "required" in all_of_rule["then"]:
                if "aggregation_group" in all_of_rule["then"]["properties"]:
                    prop = all_of_rule["then"]["properties"]["aggregation_group"]
                    if "const" in prop and prop["const"] == group_name:
                        return True
                    if "pattern" in prop and group_name in prop["pattern"]:
                        return True
    return False


def check_aggregation_group_multiple(sheet_aggr_group, schema, field_name):
    errors = []

    schema_group_names = get_schema_property(schema, field_name, "enum")
    if not schema_group_names:
        errors.append(
            f"    Error: {field_name} - Cannot find aggregation group in json schema"
        )
        return errors

    sheet_group_names = None
    if sheet_aggr_group:
        sheet_group_names = clean_aggregation_group_names(sheet_aggr_group)

    for sheet_group_name in sheet_group_names:
        in_all_of = group_in_all_of(schema, sheet_group_name)
        if sheet_group_name not in schema_group_names and not in_all_of:
            errors.append(
                f'    Error: {field_name} - Sheet has a group "{sheet_group_name}" '
                "that is missing in the schema"
            )

    for schema_group_name in schema_group_names:
        if schema_group_name not in sheet_group_names:
            errors.append(
                f'    Error: {field_name} - Schema has a group "{schema_group_name}" '
                f"that is not found in the sheet"
            )

    return errors


def verify(sheet, schema):
    # Check methods are looking for a single field
    # Compare methods run on *every* field

    errors = []
    error_check_functions = [
        compare_type,
        compare_length,
        check_required,
        check_form_type,
        check_entity_type,
        check_contribution_amount,
        check_aggregation_group,
        check_transaction_type_identifier,
    ]

    minor_errors = []
    minor_error_check_functions = [
        compare_sample_data,
    ]

    if DEBUG:
        print("    Verifying the JSON schema")

    fields = get_schema_fields(sheet)
    for field in fields.keys():
        if DEBUG:
            print(f"        {field}")
        row = sheet[str(fields[field])]
        for check_function in error_check_functions:
            if DEBUG and VERBOSE:
                print(" " * 11, check_function)
            errors += check_function(row, schema, field)
        for check_function in minor_error_check_functions:
            if DEBUG and VERBOSE:
                print(" " * 11, check_function)
            minor_errors += check_function(row, schema, field)

    return [errors, minor_errors]


def generate_report(
    errors,
    minor_errors,
    missing_schema_files,
    missing_transaction_type_identifiers,
    failed_to_open,
    failed_to_load,
    save=True,
):
    report = "\n"

    sheets_with_errors = list(errors.keys())
    sheets_with_minor_errors = list(minor_errors.keys())
    sheets_with_errors.sort()
    sheets_with_minor_errors.sort()
    missing_schema_files.sort()
    missing_transaction_type_identifiers.sort()
    failed_to_open.sort()
    failed_to_load.sort()

    if len(missing_schema_files) > 0:
        report += "Sheets without a corresponding JSON file:\n"
        report += "    " + "\n    ".join(missing_schema_files) + "\n\n"

    if len(missing_transaction_type_identifiers) > 0:
        report += "Sheets missing a Transaction Type Identifier:\n"
        report += "    " + "\n    ".join(missing_transaction_type_identifiers) + "\n\n"

    if len(failed_to_open) > 0:
        report += "Failed to open:\n"
        report += "    " + "\n    ".join(failed_to_open) + "\n\n"

    if len(failed_to_load) > 0:
        report += "Failed to load:\n"
        report += "    " + "\n    ".join(failed_to_load) + "\n\n"

    sheets = sheets_with_errors
    if display_minor_errors:
        sheets += sheets_with_minor_errors

    for sheet_name in sheets:
        if len(errors[sheet_name]) > 0:
            report += sheet_name + "\n"
            for error in errors[sheet_name]:
                report += error + "\n"
            if display_minor_errors:
                for minor_error in minor_errors[sheet_name]:
                    report += minor_error + "\n"
            report += "\n\n"

    print(report)
    if save:
        file = open("spec_verification_report.txt", "w")
        file.write(report)
        file.close()


if __name__ == "__main__":
    filename = EXCEL_FILENAME
    display_minor_errors = VERBOSE

    if not filename:
        files = listdir("./")
        xlsx_files = []
        for file in files:
            if re.search("\\.xlsx$", file):
                xlsx_files.append(file)
        if len(xlsx_files) > 0:
            xlsx_files.sort()
            filename = xlsx_files[-1]
            print("\nTesting against found spreadsheet:", filename)
            time.sleep(1.5)

    if not filename:
        print("No excel file found")
        exit()
    if not path.exists(filename):
        print("File does not exist:", filename)
        exit()

    workbook = load_workbook(filename)
    sheets = workbook._sheets

    # Sheet titles cannot be longer than 31 characters
    excluded_sheets = [
        "HDR Record",
        "TEXT",
        "zzEARMARK_MEMO_HEADQUARTERS_ACCOUNT"[:31],
        "zzEARMARK_RECEIPT_HEADQUARTERS_ACCOUNT"[:31],
        "zzEARMARK_MEMO_CONVENTION_ACCOUNT"[:31],
        "zzEARMARK_RECEIPT_CONVENTION_ACCOUNT"[:31],
        "zzEARMARK_MEMO_RECOUNT_ACCOUNT"[:31],
        "zzEARMARK_RECEIPT_RECOUNT_ACCOUNT"[:31],
    ]

    errors = {}
    minor_errors = {}
    missing_schema_files = []
    missing_transaction_type_identifiers = []
    failed_to_open = []
    failed_to_load = []

    for sheet in sheets:
        if sheet.title in excluded_sheets:
            continue

        if DEBUG:
            print(sheet.title)

        filename = get_filename(sheet)
        schema_file_path = f"../schema/{filename}.json"
        if not path.exists(schema_file_path):
            missing_schema_files.append(f"{sheet.title}: {schema_file_path}")
            continue

        json_file = open(schema_file_path, "r")
        if not json_file:
            failed_to_open.append(f"Failed to open JSON file: {filename}")
            continue

        schema = json.load(json_file)
        if not schema:
            failed_to_load.append(f"Failed to load JSON: {filename}")
            continue

        new_errors, new_minor_errors = verify(sheet, schema)
        errors[filename] = new_errors
        minor_errors[filename] = new_minor_errors

    generate_report(
        errors,
        minor_errors,
        missing_schema_files,
        missing_transaction_type_identifiers,
        failed_to_open,
        failed_to_load,
        save=True,
    )
